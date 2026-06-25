"""
acceptance_test.py
===================
Builds a realistic, multi-line-heavy bank statement with reportlab's *canvas*
(absolutely positioned text + right-aligned amounts — NOT a clean table) so it
genuinely exercises coordinate-based column assignment and multi-line
reassembly. Then runs the full pipeline and asserts the spec's acceptance
criteria (a)-(g):

    (a) each Particulars cell reads in PDF top-to-bottom order
    (b) no description contains text from a neighbouring transaction
    (c) dates align with their correct transactions
    (d) cheque numbers land in Cheque/Ref No, not inside Particulars
    (e) both date columns (Date + Value Date) are captured
    (f) the opening "Brought Forward" balance is captured and row 1 validates OK
    (g) the summary totals reconcile
"""
import os
import tempfile

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from extractor import extract_pages
from normalizer import StatementNormalizer
from validator import validate, OK
import exporter

# Column geometry (points). Left columns left-aligned; amounts right-aligned at
# their right edge so numeric header centres line up with the numbers.
X_DATE = 40
X_VDATE = 100
X_PARTIC = 175
X_CHQ = 325
R_WDR = 435     # right edge for Withdrawal
R_DEP = 505     # right edge for Deposit
R_BAL = 578     # right edge for Balance

PAGE_W, PAGE_H = A4


def build_statement(path: str) -> dict:
    """Draw the statement and return the ground-truth we will assert against."""
    c = canvas.Canvas(path, pagesize=A4)

    # --- transactions: (date, vdate, [desc lines], chq, withdrawal, deposit, balance)
    bf = 119544.23
    txns = [
        ("02/04/24", "03/04/24",
         ["NEFT DR-HDFC0001234-RAJESH",
          "KUMAR-SALARY PAYMENT FOR",
          "MARCH 2024"],
         "", 5000.00, None),
        ("05/04/24", "05/04/24",
         ["UPI-SWIGGY ORDER",
          "REF NUMBER 998877"],
         "000123", 2500.50, None),
        ("07/04/24", "08/04/24",
         ["CHEQUE DEPOSIT",
          "MUMBAI MAIN BRANCH"],
         "456789", None, 50000.00),
        ("10/04/24", "10/04/24",
         ["IMPS-AXIS BANK-VENDOR",
          "PAYMENT INVOICE 4471"],
         "", 12750.75, None),
        ("15/04/24", "16/04/24",
         ["INTEREST CREDIT QUARTERLY"],
         "", None, 1188.90),
    ]

    # Compute running balances (ground truth).
    bal = bf
    rows = []
    for d, vd, desc, chq, wdr, dep in txns:
        bal = bal - (wdr or 0) + (dep or 0)
        rows.append((d, vd, desc, chq, wdr, dep, round(bal, 2)))

    def draw_page(start_y):
        c.setFont("Helvetica-Bold", 13)
        c.drawString(X_DATE, PAGE_H - 40, "HDFC BANK LTD")
        c.setFont("Helvetica", 9)
        c.drawString(X_DATE, PAGE_H - 55, "Statement of Account   A/c No: 50100XXXXXX")

        y = PAGE_H - start_y
        # header row (numeric headers right-aligned over their numbers)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(X_DATE, y, "Date")
        c.drawString(X_VDATE, y, "Value Date")
        c.drawString(X_PARTIC, y, "Particulars")
        c.drawString(X_CHQ, y, "Chq No")
        c.drawRightString(R_WDR, y, "Withdrawal")
        c.drawRightString(R_DEP, y, "Deposit")
        c.drawRightString(R_BAL, y, "Balance")
        return y - 18

    y = draw_page(90)

    # Brought-forward row: balance only, no debit/credit.
    c.setFont("Helvetica", 8)
    c.drawString(X_DATE, y, "01/04/24")
    c.drawString(X_VDATE, y, "01/04/24")
    c.drawString(X_PARTIC, y, "BROUGHT FORWARD")
    c.drawRightString(R_BAL, y, f"{bf:,.2f}")
    y -= 16

    for d, vd, desc, chq, wdr, dep, rbal in rows:
        c.setFont("Helvetica", 8)
        c.drawString(X_DATE, y, d)
        c.drawString(X_VDATE, y, vd)
        c.drawString(X_PARTIC, y, desc[0])
        if chq:
            c.drawString(X_CHQ, y, chq)
        if wdr is not None:
            c.drawRightString(R_WDR, y, f"{wdr:,.2f}")
        if dep is not None:
            c.drawRightString(R_DEP, y, f"{dep:,.2f}")
        c.drawRightString(R_BAL, y, f"{rbal:,.2f}")
        y -= 12
        # continuation (description-only) lines
        for cont in desc[1:]:
            c.drawString(X_PARTIC, y, cont)
            y -= 12
        y -= 4

    # footer noise that must be stripped
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(X_DATE, 40, "This is a computer generated statement. Page 1 of 1")
    c.save()

    return {"bf": bf, "rows": rows, "txns": txns}


def run():
    path = os.path.join(tempfile.gettempdir(), "acceptance_statement.pdf")
    truth = build_statement(path)

    norm = StatementNormalizer()
    debug_pages = []
    for page in extract_pages(path):
        norm.feed_page(page)
        debug_pages.append((page.page_num, page.raw_debug))
    out = norm.finalize()
    validate(out, norm.opening_balance)

    failures = []

    def expect(label, cond):
        print(f"[{'PASS' if cond else 'FAIL'}] {label}")
        if not cond:
            failures.append(label)

    print("== parsed transactions ==")
    for t in out:
        print(f"  {t['date'].strftime('%d-%m-%y') if t['date'] else '----':9} "
              f"V={t['value_date'].strftime('%d-%m-%y') if t['value_date'] else '----':9} "
              f"chq={t['ref_no']!r:10} D={t['debit']} C={t['credit']} B={t['balance']} "
              f"[{t['validation']}] | {t['particulars']}")
    print()

    # Right number of transactions (5; BF is NOT a transaction).
    expect("5 transactions parsed (BF excluded)", len(out) == 5)
    if len(out) != 5:
        print("\nFAILED — wrong transaction count, aborting detailed checks.")
        return 1

    rows = truth["rows"]

    # (f) opening balance captured + first row validates OK
    expect("(f) opening BROUGHT FORWARD captured", abs((norm.opening_balance or 0) - truth["bf"]) < 0.01)
    expect("(f) first transaction validates OK", out[0]["validation"] == OK)

    # (a) particulars in correct top-to-bottom order
    expect("(a) txn0 particulars in order",
           out[0]["particulars"] == "NEFT DR-HDFC0001234-RAJESH KUMAR-SALARY PAYMENT FOR MARCH 2024")
    expect("(a) txn2 particulars in order",
           out[2]["particulars"] == "CHEQUE DEPOSIT MUMBAI MAIN BRANCH")

    # (b) no leakage between neighbours
    expect("(b) no SWIGGY leak into txn0", "SWIGGY" not in out[0]["particulars"])
    expect("(b) no SALARY leak into txn1", "SALARY" not in out[1]["particulars"])
    expect("(b) no IMPS leak into txn2", "IMPS" not in out[2]["particulars"])

    # (c) dates aligned with correct transactions
    expect("(c) txn dates align", [t["date"].strftime("%d/%m/%y") for t in out] ==
           [r[0] for r in rows])

    # (d) cheque numbers in ref column, not particulars
    expect("(d) txn1 cheque in ref column", out[1]["ref_no"].replace(" ", "") == "000123")
    expect("(d) txn1 cheque not in particulars", "000123" not in out[1]["particulars"])
    expect("(d) txn2 cheque in ref column", out[2]["ref_no"].replace(" ", "") == "456789")

    # (e) both date columns captured
    expect("(e) Date column captured", all(t["date"] is not None for t in out))
    expect("(e) Value Date column captured", all(t["value_date"] is not None for t in out))
    expect("(e) Date != Value Date where expected (txn0)",
           out[0]["date"] != out[0]["value_date"])

    # amounts split into the right columns
    expect("debit/credit split: txn0 debit", abs((out[0]["debit"] or 0) - 5000.00) < 0.01 and out[0]["credit"] is None)
    expect("debit/credit split: txn2 credit", abs((out[2]["credit"] or 0) - 50000.00) < 0.01 and out[2]["debit"] is None)

    # all rows validate OK (clean statement)
    expect("all rows validate OK", all(t["validation"] == OK for t in out))

    # (g) summary totals reconcile
    total_debit = sum(t["debit"] or 0 for t in out)
    total_credit = sum(t["credit"] or 0 for t in out)
    closing = out[-1]["balance"]
    expect("(g) opening + credits - debits == closing",
           abs((truth["bf"] + total_credit - total_debit) - closing) < 0.01)

    # export round-trips with the debug sheet
    xlsx = os.path.join(tempfile.gettempdir(), "acceptance_organized.xlsx")
    exporter.export_to_excel(out, xlsx, opening_balance=norm.opening_balance, debug_pages=debug_pages)
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    expect("Transactions sheet exists", "Transactions" in wb.sheetnames)
    expect("Debug — Raw Extract sheet exists", "Debug — Raw Extract" in wb.sheetnames)
    os.remove(xlsx)

    print()
    if failures:
        print(f"ACCEPTANCE FAILED — {len(failures)} check(s): {failures}")
        return 1
    print("ALL ACCEPTANCE CRITERIA PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
