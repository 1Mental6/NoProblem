"""Quick self-test for the parsing/normalizing/export logic (no real PDF needed).

Feeds synthetic PageData objects that mimic the messy cases described in the
spec, then checks the normalized output, validation and Excel export.
"""
from extractor import PageData
from normalizer import StatementNormalizer, clean_amount, parse_date
from validator import validate, validation_summary
import exporter, os, tempfile


def page(rows, num=1, total=1):
    return PageData(page_num=num, total_pages=total, method="test", rows=rows)


def check(label, cond):
    print(f"[{'PASS' if cond else 'FAIL'}] {label}")
    assert cond, label


# --- 1. amount cleaning -----------------------------------------------------
print("== amount cleaning ==")
check("Indian grouping 1,00,000", clean_amount("1,00,000.00")[0] == 100000.0)
check("currency Rs.", clean_amount("Rs. 5,000.50")[0] == 5000.50)
check("rupee symbol", clean_amount("₹ 12,345.67")[0] == 12345.67)
check("trailing Cr", clean_amount("2,500.00 Cr") == (2500.0, "CR"))
check("trailing Dr", clean_amount("900.00Dr") == (900.0, "DR"))
check("parentheses neg", clean_amount("(1,200.00)") == (-1200.0, "DR"))
check("blank -> None", clean_amount("-")[0] is None)

# --- 2. date parsing --------------------------------------------------------
print("== date parsing ==")
check("DD/MM/YYYY", parse_date("05/04/2023").day == 5 and parse_date("05/04/2023").month == 4)
check("DD-MMM-YYYY", parse_date("05-Apr-2023").month == 4)
check("non-date None", parse_date("5000.00") is None)

# --- 3. HDFC-style: separate Withdrawal/Deposit + multi-line narration ------
print("== HDFC-style separate columns + multi-line ==")
rows = [
    ["Date", "Narration", "Chq/Ref No", "Withdrawal Amt", "Deposit Amt", "Closing Balance"],
    ["01/04/2023", "OPENING BALANCE", "", "", "", "10,000.00"],
    ["02/04/2023", "UPI/PAYTM/123", "REF001", "1,500.00", "", "8,500.00"],
    ["", "MERCHANT GROCERY", "", "", "", ""],            # continuation row
    ["03/04/2023", "SALARY CREDIT NEFT", "REF002", "", "50,000.00", "58,500.00"],
]
n = StatementNormalizer()
n.feed_page(page(rows))
txns = n.finalize()
check("2 transactions parsed (OPENING BALANCE captured separately)", len(txns) == 2)
check("opening balance captured even though dated", n.opening_balance == 10000.0)
check("multi-line merged into particulars",
      "MERCHANT GROCERY" in txns[0]["particulars"] and "UPI/PAYTM" in txns[0]["particulars"])
check("debit split correct", txns[0]["debit"] == 1500.0 and txns[0]["credit"] is None)
check("credit split correct", txns[1]["credit"] == 50000.0)

# date-LESS opening balance line is captured as opening_balance (not a txn)
rows2 = [
    ["Date", "Narration", "Chq/Ref No", "Withdrawal Amt", "Deposit Amt", "Closing Balance"],
    ["", "Opening Balance B/F", "", "", "", "25,000.00"],     # no date -> summary capture
    ["02/04/2023", "ATM", "R1", "1,000.00", "", "24,000.00"],
]
n2 = StatementNormalizer()
n2.feed_page(page(rows2))
t2 = n2.finalize()
check("date-less opening captured separately", n2.opening_balance == 25000.0 and len(t2) == 1)

# --- 4. ICICI-style: single Amount column + Dr/Cr indicator -----------------
print("== single Amount column + Dr/Cr indicator ==")
rows = [
    ["Txn Date", "Value Date", "Transaction Remarks", "Amount", "Type", "Balance"],
    ["05/04/2023", "05/04/2023", "ATM WDL", "3,000.00", "Dr", "55,500.00"],
    ["06/04/2023", "06/04/2023", "INTEREST", "120.00", "Cr", "55,620.00"],
]
n = StatementNormalizer()
n.feed_page(page(rows))
txns = n.finalize()
check("amount->debit via indicator", txns[0]["debit"] == 3000.0 and txns[0]["credit"] is None)
check("amount->credit via indicator", txns[1]["credit"] == 120.0 and txns[1]["debit"] is None)
check("value_date kept separate", txns[0]["value_date"] is not None)

# --- 5. fuzzy header variants + repeated header on page 2 -------------------
print("== fuzzy headers + repeated header across pages ==")
p1 = [
    ["Tran Date", "Particulars", "Ref No.", "Debit", "Credit", "Running Balance"],
    ["01/05/2023", "OPENING BAL", "", "", "", "1,00,000.00"],
    ["02/05/2023", "NEFT OUT", "N1", "10,000.00", "", "90,000.00"],
]
p2 = [
    ["Tran Date", "Particulars", "Ref No.", "Debit", "Credit", "Running Balance"],  # repeated header
    ["03/05/2023", "CHEQUE DEP", "C9", "", "5,000.00", "95,000.00"],
    ["This is a computer generated statement", "", "", "", "", ""],   # footer noise
]
n = StatementNormalizer()
n.feed_page(page(p1, 1, 2))
n.feed_page(page(p2, 2, 2))
txns = n.finalize()
check("fuzzy header mapped (Tran Date/Running Balance)", len(txns) == 3)
check("repeated header on page 2 stripped", all(t["date"] is not None for t in txns))
check("footer noise not a transaction", txns[-1]["particulars"] == "CHEQUE DEP")

# --- 6. validation ----------------------------------------------------------
print("== validation ==")
validate(txns, opening_balance=100000.0)
# 100000 (open) -10000 = 90000 OK ; 90000 +5000 = 95000 OK
check("all rows validate OK", validation_summary(txns)["checks"] == 0)
# inject an error in the LAST row (a mid-row error would cascade to the next)
bad = [dict(t) for t in txns]
bad[-1]["balance"] = 99999.0
validate(bad, opening_balance=100000.0)
check("CHECK flag detected", validation_summary(bad)["checks"] == 1)

# --- 7. OCR text-line mode --------------------------------------------------
print("== OCR text-line fallback ==")
lines = [
    "01/06/2023 ATM CASH WITHDRAWAL 2,000.00 48,000.00",
    "extra narration line for above",
    "02/06/2023 SALARY 30,000.00 78,000.00",
]
n = StatementNormalizer()
n.opening_balance = 50000.0
n._last_balance = 50000.0
n.feed_page(PageData(1, 1, "ocr", text_lines=lines, ocr_used=True))
txns = n.finalize()
check("OCR parsed 2 txns", len(txns) == 2)
check("OCR continuation merged", "extra narration" in txns[0]["particulars"])
check("OCR inferred debit (balance fell)", txns[0]["debit"] == 2000.0)
check("OCR inferred credit (balance rose)", txns[1]["credit"] == 30000.0)
check("ocr_used flag set", n.ocr_used is True)

# --- 8. Excel export --------------------------------------------------------
print("== Excel export ==")
validate(txns, opening_balance=50000.0)
out = os.path.join(tempfile.gettempdir(), "selftest_organized.xlsx")
exporter.export_to_excel(txns, out, opening_balance=50000.0, ocr_used=True)
from openpyxl import load_workbook
wb = load_workbook(out)
ws = wb["Transactions"]
check("xlsx file written", os.path.exists(out))
check("header present", [c.value for c in ws[2]][:3] == ["Date", "Value Date", "Particulars"] or
      [c.value for c in ws[1]][:3] == ["Date", "Value Date", "Particulars"])
check("summary block written", any("Closing balance" == (row[0].value)
      for row in ws.iter_rows() if row and row[0].value))
os.remove(out)

print("\nALL SELF-TESTS PASSED")
