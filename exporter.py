"""
exporter.py
===========
Write the normalized + validated transactions to a polished .xlsx (openpyxl).

Sheet 1 — "Transactions"
    Columns: Date | Value Date | Particulars | Cheque/Ref No | Debit | Credit
             | Balance | Validation
    * Bold, filled, frozen header row; auto-width columns.
    * Debit / Credit / Balance right-aligned and number-formatted (#,##0.00).
    * Rows flagged "CHECK" are highlighted so errors jump out.
    * A summary block at the bottom: opening balance, total debits, total
      credits, closing balance, transaction count.
    * Over Excel's ~1,048,576-row limit -> split across multiple sheets.

Sheet 2 — "Debug — Raw Extract"
    The raw extracted lines (with x/y coordinates) exactly as pdfplumber read
    them, labelled per page. Lets the user diagnose alignment on odd statements
    by comparing the raw extract against the final output.

Default filename: <original_pdf_name>_organized.xlsx
"""

from __future__ import annotations

import os
from typing import List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["Date", "Value Date", "Particulars", "Cheque/Ref No",
           "Debit", "Credit", "Balance", "Validation"]
_FIELDS = ["date", "value_date", "particulars", "ref_no", "debit", "credit", "balance", "validation"]

_DATE_FORMAT = "%d-%m-%Y"
_NUMBER_FORMAT = "#,##0.00"
_MAX_ROWS_PER_SHEET = 1_000_000
_DEBUG_SHEET = "Debug — Raw Extract"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUMMARY_FONT = Font(bold=True)
_CHECK_FILL = PatternFill("solid", fgColor="FFF2CC")
_PAGE_FONT = Font(bold=True, color="1F4E78")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_MONEY_COLS = {5, 6, 7}   # 1-based: Debit, Credit, Balance


def default_output_path(pdf_path: str) -> str:
    """`/path/foo.pdf` -> `/path/foo_organized.xlsx`."""
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    folder = os.path.dirname(os.path.abspath(pdf_path))
    return os.path.join(folder, f"{base}_organized.xlsx")


def export_to_excel(
    transactions: List[dict],
    output_path: str,
    opening_balance: Optional[float] = None,
    ocr_used: bool = False,
    debug_pages: Optional[Sequence[Tuple[int, Sequence[str]]]] = None,
    unrecognized: bool = False,
) -> str:
    """Write *transactions* (+ optional debug dump) to *output_path*.

    *unrecognized* writes a prominent banner warning that the bank format was not
    recognised and the rows were parsed with the best-effort generic parser.
    """
    wb = Workbook()
    wb.remove(wb.active)

    chunks = _chunk(transactions, _MAX_ROWS_PER_SHEET) or [[]]
    multi = len(chunks) > 1

    last_ws = None
    for i, chunk in enumerate(chunks, start=1):
        title = f"Transactions {i}" if multi else "Transactions"
        last_ws = _write_sheet(wb, title, chunk, ocr_used, unrecognized)

    _write_summary(last_ws, transactions, opening_balance)
    _write_debug_sheet(wb, debug_pages or [])

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Transactions sheet
# ---------------------------------------------------------------------------
def _write_sheet(wb: Workbook, title: str, transactions: List[dict],
                 ocr_used: bool, unrecognized: bool = False):
    ws = wb.create_sheet(title=title)
    row_cursor = 1

    for note in _banners(ocr_used, unrecognized):
        ws.cell(row=row_cursor, column=1, value=note).font = Font(bold=True, color="C00000")
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=len(HEADERS))
        row_cursor += 1

    header_row = row_cursor
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    for txn in transactions:
        row_cursor += 1
        is_check = txn.get("validation") == "CHECK"
        for col, field in enumerate(_FIELDS, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=_format_value(field, txn.get(field)))
            cell.border = _BORDER
            if col in _MONEY_COLS:
                cell.number_format = _NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
            if is_check:
                cell.fill = _CHECK_FILL

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autofit(ws, header_row)
    return ws


def _banners(ocr_used: bool, unrecognized: bool) -> List[str]:
    """Warning lines to show above the header (most severe first)."""
    notes = []
    if unrecognized:
        notes.append("⚠ Bank format NOT recognised — parsed with a best-effort generic "
                     "parser. Review every row carefully before relying on this data.")
    if ocr_used:
        notes.append("⚠ Some pages were read with OCR (scanned PDF). "
                     "Please review the extracted figures manually.")
    return notes


def _write_summary(ws, transactions: List[dict], opening_balance: Optional[float]) -> None:
    total_debit = sum(t.get("debit") or 0.0 for t in transactions)
    total_credit = sum(t.get("credit") or 0.0 for t in transactions)
    balances = [t.get("balance") for t in transactions if t.get("balance") is not None]
    opening = opening_balance if opening_balance is not None else (balances[0] if balances else None)
    closing = balances[-1] if balances else None
    checks = sum(1 for t in transactions if t.get("validation") == "CHECK")

    start = (ws.max_row or 1) + 2
    rows = [
        ("Summary", None),
        ("Opening balance", opening),
        ("Total debits", total_debit),
        ("Total credits", total_credit),
        ("Closing balance", closing),
        ("Transaction count", len(transactions)),
        ("Validation issues (CHECK)", checks),
    ]
    for i, (label, value) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=label).font = _SUMMARY_FONT
        if value is not None:
            vc = ws.cell(row=r, column=2, value=value)
            if isinstance(value, float):
                vc.number_format = _NUMBER_FORMAT
                vc.alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------------
# Debug sheet
# ---------------------------------------------------------------------------
def _write_debug_sheet(wb: Workbook, debug_pages: Sequence[Tuple[int, Sequence[str]]]) -> None:
    ws = wb.create_sheet(title=_DEBUG_SHEET)
    intro = ("Raw extract — each word is shown with its [x=left, y=top] coordinate, "
             "exactly as read from the PDF. Compare against the Transactions sheet "
             "to diagnose column-alignment issues on unusual layouts.")
    ws.cell(row=1, column=1, value=intro).font = Font(italic=True, color="808080")

    r = 3
    for page_num, lines in debug_pages:
        ws.cell(row=r, column=1, value=f"───── Page {page_num} ─────").font = _PAGE_FONT
        r += 1
        for line in lines:
            ws.cell(row=r, column=1, value=line)
            r += 1
        r += 1   # blank spacer between pages
    if not debug_pages:
        ws.cell(row=3, column=1, value="(No raw-coordinate data — page(s) came from OCR or camelot.)")

    ws.column_dimensions["A"].width = 150


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _format_value(field: str, value):
    if value is None:
        return None
    if field in ("date", "value_date"):
        try:
            return value.strftime(_DATE_FORMAT)
        except AttributeError:
            return str(value)
    if field in ("debit", "credit", "balance"):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    return value


def _autofit(ws, header_row: int) -> None:
    widths = {}
    for row in ws.iter_rows(min_row=header_row):
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            widths[col] = max(widths.get(col, 0), len(str(cell.value)))
    for col, length in widths.items():
        cap = 60 if col == 3 else 28   # Particulars (col 3) can be long
        ws.column_dimensions[get_column_letter(col)].width = min(max(length + 2, 10), cap)


def _chunk(items: List[dict], size: int) -> List[List[dict]]:
    if not items:
        return []
    return [items[i:i + size] for i in range(0, len(items), size)]
